
from autosectask.utility.util_logging import SimpleLogger
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension


logger = SimpleLogger().get_logger()


class XlsxFS:

    def __init__(self, path):
        self.__path = path
        self.__wb = load_workbook(path)

    def __enter__(self):
        logger.info(f"opening xlsx file {self.__path}")
        return self

    def __exit__(self, e_t, e_v, t_b):
        logger.info("close xlsx file...")
        self.__wb.save(self.__path)
        self.__wb.close()

    def create_sheet(self, sheet_name):
        sheet =self.__wb.create_sheet(sheet_name)
        return sheet

    def write_column(self, sheet_name, contents, title=None):
        if title:
            contents = title.splitlines() + contents
        sheet = self.__wb[sheet_name]
        for i in range(len(contents)):
            index = "A"+str(i+1)
            logger.debug(f"更新A列，内容：{contents[i]}")
            sheet[index] = contents[i]

    def write_cells(self, sheet_name, cells, content):
        contents = content if isinstance(content, list) else [content for _ in range(0, len(cells))]
        sheet = self.__wb[sheet_name]
        logger.info(sheet)
        for cell, content in zip(cells, contents):
            logger.info(f"更新单元格：{cell}，内容：{content}")
            sheet[cell] = content
        # self.__wb.save(self.__path)

    def set_cell_format(self, sheet_name, cell_index, width=None, height=None, background=None,
                        font=None, alignment=None):
        sheet = self.__wb[sheet_name]
        cell = sheet[cell_index]
        # 设置单元格的宽度
        if width:
            column_dimension = ColumnDimension(sheet, min=cell.column, max=cell.column, width=width)
            sheet.column_dimensions[get_column_letter(cell.column)] = column_dimension

        # 设置单元格的高度
        if height:
            row_dimension = RowDimension(sheet, min=cell.row, max=cell.row, height=height)
            sheet.row_dimensions[cell.row] = row_dimension

        # 设置单元格的背景色
        if background:
            cell.fill = background

        # 设置单元格的字体及颜色
        if font:
            cell.font = font

        # 设置单元格的对齐方式
        if alignment:
            cell.alignment = alignment

    def insert_imgs(self, sheet_name, cells, imgs):
        sheet = self.__wb[sheet_name]
        logger.info(sheet)
        for cell, image in zip(cells, imgs):
            logger.info(f"插入图片：{cell}, {image}")
            img = Image(image)
            img.width, img.height = 205, 130
            sheet.add_image(img, cell)
        # self.__wb.save(self.__path)

    def del_rows(self, sheet_name, start, end):
        sheet = self.__wb[sheet_name]
        logger.info(f"删除{str(sheet)}行范围：[{start},{end}]")
        sheet.delete_rows(start, end)
        # self.__wb.save(self.__path)

    def reset_rows_value(self, sheet_name, start, end, columns):
        sheet = self.__wb[sheet_name]
        for column in columns:
            logger.info(f"重置{str(sheet)}列{column}, [{start},{end}]内容为空")
            for i in range(start, end + 1):
                sheet[f'{column}{i}'] = ""
        # self.__wb.save(self.__path)
