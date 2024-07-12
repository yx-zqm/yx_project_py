# SMP巡检报告自动生成
# !!!!! 把get_server_log作为一个service，通过controller来拼接业务逻辑
import os

from openpyxl.styles import PatternFill

from autosectask.dao.fs_smp_routing_config import RoutingConfigDAO
from autosectask.utility.util_conn_ssh import SSHConnection
from autosectask.dao.fs_template import TemplateDAO
from autosectask.dao.fs_xlsx import XlsxFS


if __name__ == '__main__':
    # 登录服务器获取日志内容
    a = r"""C:\Users\KowkaNonk\Documents\OneDrive\AsiaSec\Workspace\sh4a\ais\autosectask\config\smp_routing_inspection.ini"""
    cmd = r"""cd /opt/mcb/smp/program/smc-dataupload-5.1/logs && grep "上报完成" ./du_2024-06*"""
    cfg = RoutingConfigDAO(a)
    host = RoutingConfigDAO(a).get_server_cfg()
    with SSHConnection(ip=host["ip"], username=host["user"], password=host["pp"]) as ssh:
        # 执行命令
        result = ssh.execute_command(cmd)

        # 复制模板到新的目录 output_file
        output_file_pat = cfg.get_output_file()
        output_file = os.path.join(cfg.get_output_path(), output_file_pat)
        TemplateDAO.copy_template(cfg.get_template_path(), output_file)

        # 写入文件
        with XlsxFS(output_file) as route_xlsx:
            sheet = route_xlsx.create_sheet(cfg.get_sheet_name())
            route_xlsx.write_column(cfg.get_sheet_name(), result.splitlines(), title="SMC考核上报日志")
            # 设置单元格
            bg_fill = PatternFill(fill_type='solid', fgColor='FFFF00')
            route_xlsx.set_cell_format(cfg.get_sheet_name(), "A1", width=150, background=bg_fill)
